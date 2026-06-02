import re
from pathlib import Path
from openpyxl import load_workbook
import pandas as pd
from logger import escribir_detalle
from openpyxl.utils import get_column_letter



# =====================================================
# REGEX
# =====================================================

# ROMAN_RE = re.compile(r"^[IVXLCDM]+\.", re.I)
# NUM_RE = re.compile(r"^\d+\.")
# LETTER_RE = re.compile(r"^[a-z]\.", re.I)


# =====================================================
# UTILIDADES
# =====================================================

def limpiar(texto):
    if texto is None:
        return ""
    
    texto = str(texto).strip()
    texto = re.sub(r"\s+", " ", texto)
    return texto


def es_numero(valor):
    if valor is None:
        return False

    try:
        float(str(valor).replace(",", ""))
        return True
    except:
        return False


def obtener_pn(cell):
    try:
        if cell.hyperlink is None:
            return None

        target = cell.hyperlink.target
        if not target:
            return None

        m = PN_RE.search(target)
        if m:
            return m.group(1)
    except:
        pass
    return None


def tiene_hyperlink(cell):
    try:
        return cell.hyperlink is not None
    except:
        return False


def es_continuacion(ws, fila, col_desc):
    cell = ws.cell(fila, col_desc)
    texto = limpiar(cell.value)

    if not texto:
        return False

    texto_upper = texto.upper()
    # NOTAS nunca son continuación
    if texto_upper.startswith("NOTA"):
        return False

    # Las unidades deben quedar como hijos
    if texto.startswith("("):
        return False

    # Si tiene PN es una descripción propia
    if obtener_pn(cell):
        return False

    # Si tiene hyperlink es descripción propia
    if tiene_hyperlink(cell):
        return False

    # Si tiene datos numéricos asociados
    if tiene_datos(ws, fila, col_desc):
        return False

    nivel = obtener_nivel(texto)

    # Nueva jerarquía real
    if nivel in (1, 2, 3, 4):
        return False

    return True


def detectar_columna_descripciones(ws):

    mejor_col = None
    mejor_score = -1

    for col in range(
        1,
        min(ws.max_column, 15) + 1
    ):

        textos = 0
        numeros = 0
        filas_con_datos = 0

        for fila in range(
            1,
            ws.max_row + 1
        ):

            valor = ws.cell(
                fila,
                col
            ).value

            if valor is None:
                continue

            valor = str(valor).strip()

            if not valor:
                continue

            # Texto
            if not es_numero(valor):
                textos += 1

                # descripción con datos a la derecha
                if tiene_datos(
                    ws,
                    fila,
                    col
                ):
                    filas_con_datos += 1

            else:
                numeros += 1

        # Fórmula de puntuación
        score = (
            textos * 2
            + filas_con_datos * 20
            - numeros
        )

        # print(
        #     f"COL={get_column_letter(col)} "
        #     f"TXT={textos} "
        #     f"NUM={numeros} "
        #     f"DATOS={filas_con_datos} "
        #     f"SCORE={score}"
        # )

        if score > mejor_score:
            mejor_score = score
            mejor_col = col

    print(
        f"Columna elegida: "
        f"{get_column_letter(mejor_col)}"
    )

    return mejor_col

def construir_descripciones(ws, col_desc):
    descripciones = []
    stack = []
    filas_consumidas = set()

    for fila in range(1, ws.max_row + 1):
        if fila in filas_consumidas:
            continue

        texto = limpiar(
            ws.cell(fila, col_desc).value
        )
        if not texto:
            continue

        texto_completo = obtener_descripcion_completa(ws, fila, col_desc, filas_consumidas)
        descripcion = construir_jerarquia(texto_completo, stack)
        pn = obtener_pn(ws.cell(fila, col_desc))
        tiene_data = tiene_datos(ws, fila, col_desc)

        descripciones.append({
            "fila": fila,
            "pn": pn,
            "tiene_datos": tiene_data,
            "descripcion": descripcion
        })
    return descripciones

def obtener_descripcion_completa(ws, fila, col_desc, filas_consumidas, texto_base=None):
    if texto_base:
        partes = [texto_base]
    else:
        partes = [
            limpiar(ws.cell(fila, col_desc).value)
        ]

    fila_actual = fila + 1
    while fila_actual <= ws.max_row:
        cell_sig = ws.cell(fila_actual, col_desc)
        texto = limpiar(cell_sig.value)

        if not texto:
            fila_actual += 1
            continue

        # ==========================================
        # CASO:
        # II. ACTIVOS EXTERNOS NETOS
        # DE LARGO PLAZO (con PN)
        # ==========================================
        #
        # Se concatena para construir el encabezado
        # pero NO se consume la fila porque debe
        # generar su propio registro después.
        #
        # ==========================================

        if es_continuacion_encabezado(
            ws,
            fila_actual,
            col_desc
        ):

            partes.append(texto)

            fila_actual += 1

            continue

        # ==========================================
        # SI LA FILA TIENE PN
        # ES UN NUEVO REGISTRO
        # ==========================================

        if obtener_pn(cell_sig):
            break

        # ==========================================
        # SI TIENE DATOS
        # ES UN NUEVO REGISTRO
        # ==========================================

        if tiene_datos(
            ws,
            fila_actual,
            col_desc
        ):
            break

        # ==========================================
        # NUEVA JERARQUÍA
        # ==========================================

        nivel = obtener_nivel(texto)

        if nivel != 5:
            break

        partes.append(texto)

        filas_consumidas.add(
            fila_actual
        )

        fila_actual += 1

    descripcion = " ".join(partes)

    descripcion = re.sub(
        r"\bNota:\s*",
        "",
        descripcion,
        flags=re.I
    )

    descripcion = re.sub(
        r"\s+",
        " ",
        descripcion
    ).strip()

    return descripcion

PN_RE = re.compile(r"(PN[A-Z0-9]+)", re.I)
ROMAN_MAYOR_RE = re.compile(r'^[IVXLCDM]+\.',re.ASCII)
# NUM_RE = re.compile(r'^\d+(\.| )')
NUM_RE = re.compile(r'^\d+')
LETTER_RE = re.compile(r'^[a-z]\.')
ROMAN_MINOR_RE = re.compile(r'^(i|ii|iii|iv|v|vi|vii|viii|ix|x)\.',re.I)
ASTERISCO_RE = re.compile(r'^\*')

def obtener_nivel(texto):

    texto = texto.strip()

    if ROMAN_MAYOR_RE.match(texto):
        return 1

    if NUM_RE.match(texto):
        return 2

    # IMPORTANTE:
    # primero romanos minúsculos

    if ROMAN_MINOR_RE.match(texto):
        return 4

    if LETTER_RE.match(texto):
        return 3

    if texto.startswith("-"):
        return 5

    if texto.startswith("*"):
        return 6

    return 99


def construir_jerarquia(texto, stack):
    nivel = obtener_nivel(texto)

    if nivel == 99:
        if stack:
            return (
                " > ".join(
                    item["texto"]
                    for item in stack
                )
                + " > "
                + texto
            )
        return texto

    while stack and stack[-1]["nivel"] >= nivel:
        stack.pop()

    stack.append({
        "nivel": nivel,
        "texto": texto
    })
    return " > ".join(
        item["texto"]
        for item in stack
    )

def obtener_descripcion_hacia_arriba(ws,fila,col_desc):

    texto_actual = limpiar(
        ws.cell(fila,col_desc).value
    )

    if not texto_actual:
        return ""

    fila_arriba = fila - 1

    while fila_arriba >= 1:

        texto_arriba = limpiar(
            ws.cell(fila_arriba,col_desc).value
        )

        if not texto_arriba:
            fila_arriba -= 1
            continue

        nivel_arriba = obtener_nivel(
            texto_arriba
        )

        # =====================================
        # CASO:
        # II. ACTIVOS EXTERNOS NETOS
        # DE LARGO PLAZO
        # =====================================

        if (
            nivel_arriba == 1
            and obtener_nivel(texto_actual) == 99
        ):
            return (
                texto_arriba
                + " "
                + texto_actual
            )

        # =====================================
        # CASO:
        # I. ACTIVOS ...
        # (Millones de USD)
        # =====================================

        if (
            texto_actual.startswith("(")
            and nivel_arriba == 1
        ):
            return (
                texto_arriba
                + " "
                + texto_actual
            )

        break

    return texto_actual


def es_continuacion_encabezado(ws,fila,col_desc):

    if fila <= 1:
        return False

    texto = limpiar(
        ws.cell(fila,col_desc).value
    )

    if not texto:
        return False

    fila_arriba = fila - 1

    texto_arriba = limpiar(
        ws.cell(fila_arriba,col_desc).value
    )

    if not texto_arriba:
        return False

    nivel_actual = obtener_nivel(texto)
    nivel_arriba = obtener_nivel(texto_arriba)

    pn_actual = obtener_pn(
        ws.cell(fila,col_desc)
    )

    datos_actual = tiene_datos(
        ws,
        fila,
        col_desc
    )

    pn_arriba = obtener_pn(
        ws.cell(fila_arriba,col_desc)
    )

    datos_arriba = tiene_datos(
        ws,
        fila_arriba,
        col_desc
    )

    # =====================================
    # CASO 1
    #
    # II. ACTIVOS EXTERNOS NETOS
    # DE LARGO PLAZO
    # =====================================

    if (
        nivel_actual == 99
        and nivel_arriba == 1
        and not pn_actual
        and not datos_actual
    ):
        return True

    # =====================================
    # CASO 2
    #
    # I. ACTIVOS EXTERNOS ...
    # (Millones de USD)
    # =====================================

    if (
        texto.startswith("(")
        and nivel_arriba == 1
        and pn_arriba
    ):
        return True

    return False

# =====================================================
# VALIDACION
# =====================================================

def tiene_datos(ws, fila, col_desc):
    encontrados = 0
    for col in range(col_desc + 1, ws.max_column + 1):
        valor = ws.cell(fila, col).value

        if es_numero(valor):
            encontrados += 1

        if encontrados >= 3:
            return True 
    return False


# =====================================================
# PROCESAR CUADRO HORIZONTAL
# =====================================================

def procesar_horizontal(ruta_excel):
    wb = load_workbook(ruta_excel, data_only=False)
    ws = wb[wb.sheetnames[0]]

    numero_cuadro = re.search(
        r"(\d+)",
        Path(ruta_excel).stem
    ).group(1).zfill(3)

    resultados = []
    stack = []
    filas_consumidas = set()

    ultima_descripcion_nivel1 = None
    ultima_descripcion_real = None
    
    col_desc = detectar_columna_descripciones(ws)

    if not col_desc:
        raise Exception("No se pudo detectar columna de descripciones")

    print(
        f"Columna detectada: "
        f"{get_column_letter(col_desc)}"
    )

    escribir_detalle(
        f"CUADRO={numero_cuadro} | "
        f"COLUMNA_DESC={get_column_letter(col_desc)}"
    )

    print(f"\nProcesando cuadro {numero_cuadro}")

    for fila in range(1, ws.max_row + 1):
        if fila in filas_consumidas:
            continue

        cell = ws.cell(fila, col_desc)
        texto = limpiar(cell.value)
        if not texto:
            continue

        if texto.upper().startswith("NOTA"):
            continue

        # ==================================
        # ARMAR DESCRIPCIÓN COMPLETA
        # ==================================
        texto_completo = obtener_descripcion_hacia_arriba(ws, fila, col_desc)
        texto_completo = obtener_descripcion_completa(
            ws,
            fila,
            col_desc,
            filas_consumidas,
            texto_base=texto_completo
        )

        # ==================================
        # DETECTAR NUEVO NIVEL 1
        # ==================================
        nivel_actual = obtener_nivel(texto)

        if nivel_actual == 1:
            descripcion_tmp = construir_jerarquia(texto_completo, stack)
            ultima_descripcion_nivel1 = descripcion_tmp
            ultima_descripcion_real = descripcion_tmp
            descripcion = descripcion_tmp

        # ==================================
        # FILAS TIPO "(Millones de USD)"
        # ==================================
        elif (texto.startswith("(") and ultima_descripcion_real):
            descripcion = (
                ultima_descripcion_real
                + " > "
                + texto_completo
            )
        else:
            descripcion = construir_jerarquia(texto_completo, stack)
            ultima_descripcion_real = descripcion

        # ==================================
        # VALIDAR AL FINAL
        # ==================================
        tiene_serie = tiene_datos(ws, fila, col_desc)
        pn = obtener_pn(cell)
        if not tiene_serie and not pn:
            continue

        print(
            f"VALIDACION -> "
            f"FILA={fila} "
            f"PN={pn} "
            f"SERIE={tiene_serie}"
        )

        resultados.append({
            "num_cuadro": f"cuadro-{numero_cuadro}",
            "codigo_serie": pn,
            "descripcion": descripcion,
        })

        escribir_detalle(
            f"CUADRO={numero_cuadro} | "
            f"FILA={fila} | "
            f"PN={pn if pn else '-'} | "
            f"DESC={descripcion}"
        )

        print(
            f"Fila={fila:<4} "
            f"PN={pn if pn else '-'} "
            f"{descripcion}"
        )

    print(
        f"\nTotal encontrados: {len(resultados)}"
    )
    return resultados

# =====================================================
# PROCESAR TODOS LOS HORIZONTALES
# =====================================================

def procesar_carpeta():
    carpeta = Path("Nota_cuadros")
    resultados = []
    
    for archivo in carpeta.glob("*.xlsx"):
        print("\n" + "=" * 80)
        print(f"Procesando: {archivo.name}")
        print("=" * 80)
        try:
            datos = procesar_horizontal(archivo)
            resultados.extend(datos)
        except Exception as e:
            print(
                f"ERROR {archivo.name}: {e}"
            )

    df = pd.DataFrame(resultados)
    df.to_excel(
        "resultado_horizontal.xlsx",
        index=False
    )
    print(
        f"\nTotal registros: {len(df)}"
    )


# =====================================================
# MENU
# =====================================================

if __name__ == "__main__":
    print("\n1. Procesar cuadro-001")
    print("2. Procesar todos los horizontales")

    opcion = input("\nOpción: ").strip()
    if opcion == "1":
        datos = procesar_horizontal("cuadro-001.xlsx")
        df = pd.DataFrame(datos)
        df.to_excel("resultado_prueba.xlsx",index=False)
    elif opcion == "2":
        procesar_carpeta()