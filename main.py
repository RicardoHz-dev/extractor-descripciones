from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
import pandas as pd
import re

from extractor_horizontal import procesar_horizontal
from logger import escribir_log

CUADROS_VERTICALES = {
    "025","030","031","032","033","035","036","037","038","039","040",
    "041","042","062","072","073","083","084","112","029","017","125",
    "126","074","002","003","006","007","008","012","013","010","022",
    "026","027","043","044","045","046","047","048"
}

CUADROS_CRUZADOS = {
    "004","005","014","015","016","018","019","020",
    "034","051","052","053","009"
}

def parsear_cuadros(texto):
    cuadros = set()

    for parte in texto.split(","):

        parte = parte.strip()

        if "-" in parte:

            inicio, fin = parte.split("-")

            for n in range(int(inicio), int(fin) + 1):
                cuadros.add(str(n).zfill(3))

        else:
            cuadros.add(str(int(parte)).zfill(3))

    return sorted(cuadros)

def obtener_archivos():

    carpeta = Path("Nota_cuadros")

    archivos = {}

    for archivo in carpeta.glob("*.xlsx"):

        match = re.search(r"(\d+)", archivo.stem)

        if match:

            numero = match.group(1).zfill(3)

            archivos[numero] = archivo

    return archivos

def procesar_lista(cuadros):

    archivos = obtener_archivos()

    resultados = []

    for i, numero in enumerate(cuadros, start=1):

        print(f"\n[{i}/{len(cuadros)}] Cuadro {numero}")

        escribir_log(f"[{numero}] Inicio")

        if numero in CUADROS_VERTICALES:

            print(" -> Vertical (omitido)")
            escribir_log(f"[{numero}] Vertical - OMITIDO")
            continue

        if numero in CUADROS_CRUZADOS:

            print(" -> Cruzado (omitido)")
            escribir_log(f"[{numero}] Cruzado - OMITIDO")
            continue

        archivo = archivos.get(numero)

        if not archivo:

            print(" -> Archivo no encontrado")
            escribir_log(f"[{numero}] Archivo no encontrado")
            continue

        try:

            print(" -> Horizontal")

            datos = procesar_horizontal(
                str(archivo)
            )

            resultados.extend(datos)

            escribir_log(
                f"[{numero}] OK - {len(datos)} registros"
            )

        except Exception as e:

            print(f"ERROR: {e}")

            escribir_log(
                f"[{numero}] ERROR: {str(e)}"
            )

    return resultados

def crear_excel_resultado(filas: list[dict], ruta_salida: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparacion"

    fuente_header = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    fuente_normal = Font(name="Calibri", size=11)
    borde_fino = Border(left=Side(style="thin", color="D3D3D3"), right=Side(style="thin", color="D3D3D3"),
                        top=Side(style="thin", color="D3D3D3"), bottom=Side(style="thin", color="D3D3D3"))
    relleno_header = PatternFill("solid", fgColor="1A237E")

    anchos = {"A": 15, "B": 18, "C": 50, "D": 35, "E": 55}
    for col, w in anchos.items(): ws.column_dimensions[col].width = w

    encabezados = ["N° Cuadro", "Codigo de Serie", "Descripcion NS", "Categoria", "Nombre de la Serie"]
    for idx, texto in enumerate(encabezados, 1):
        c = ws.cell(row=1, column=idx, value=texto)
        c.font, c.fill, c.border = fuente_header, relleno_header, borde_fino
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28
    f_idx = 2

    for item in filas:
        datos = [
            item.get("num_cuadro", item.get("N_CUADRO", "")), 
            item.get("codigo_serie", item.get("CODIGO_SERIE", "")), 
            item.get("descripcion", item.get("DESCRIPCION", "")), 
            # item.get("bcrp_categoria", ""), 
            # item.get("bcrp_nombre", "")
        ]
        for col_idx, val in enumerate(datos, 1):
            celda = ws.cell(row=f_idx, column=col_idx, value=val)
            celda.font, celda.border = fuente_normal, borde_fino
            celda.alignment = Alignment(horizontal="center" if col_idx in [1,2] else "left", vertical="top", wrap_text=True)
        f_idx += 1

    ws.views.sheetView[0].showGridLines = True
    ws.freeze_panes = "A2"
    wb.save(ruta_salida)

def guardar_resultado(resultados, nombre_archivo):

    if not resultados:
        print("\nNo hay datos para exportar")
        return

    Path("output").mkdir(exist_ok=True)

    ruta = Path("output") / nombre_archivo

    crear_excel_resultado(resultados, str(ruta))

    print(f"\nArchivo generado: {ruta}")
    print(f"Registros: {len(resultados)}")

    escribir_log(
        f"Archivo generado: {ruta}"
    )

    escribir_log(
        f"Total registros: {len(resultados)}"
    )

def procesar_todos():

    cuadros = [
        str(i).zfill(3)
        for i in range(1, 127)
    ]

    resultados = procesar_lista(cuadros)

    guardar_resultado(
        resultados,
        "resultado_todos.xlsx"
    )

def procesar_especificos():

    entrada = input(
        "\nIngrese cuadros (ej: 1,5,10-20,40): "
    )

    cuadros = parsear_cuadros(
        entrada
    )

    resultados = procesar_lista(
        cuadros
    )

    guardar_resultado(
        resultados,
        "resultado_personalizado.xlsx"
    )

def mostrar_menu():

    print("\n" + "=" * 50)
    print("EXTRACTOR DE DESCRIPCIONES")
    print("=" * 50)
    print("1. Procesar TODOS")
    print("2. Procesar ESPECIFICOS")
    print("3. Salir")

    return input("\nSeleccione: ").strip()

def main():

    escribir_log("=" * 60)
    escribir_log(f"INICIO: {datetime.now()}")
    escribir_log("=" * 60)

    while True:

        opcion = mostrar_menu()

        if opcion == "1":
            procesar_todos()

        elif opcion == "2":
            procesar_especificos()

        elif opcion == "3":

            escribir_log("=" * 60)
            escribir_log(f"FIN: {datetime.now()}")
            escribir_log("=" * 60)

            print("\nFin del programa")
            break

        else:
            print("\nOpción inválida")

if __name__ == "__main__":
    main()